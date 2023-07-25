import json
import psycopg2
import openpyxl
import pandas as pd
from psycopg2.extensions import AsIs
from openpyxl import load_workbook
book = load_workbook('devicedata.xlsx')
sheet = book['Sheet1']
import datetime

DB_NAME = "devices"
DB_USER = "devices"
DB_PASS = "devicesdbtest"
DB_HOST = "localhost"
DB_PORT = "5432"
PATH = "devicedata.xlsx"
def lambda_handler():
    # TODO implement
    try:
        conn = psycopg2.connect(database=DB_NAME,
                                user=DB_USER,
                                password=DB_PASS,
                                host=DB_HOST,
                                port=DB_PORT)
        print("Database connected successfully")
        cur = conn.cursor()
        # cur.execute("SELECT * FROM  %(table)s", {"table": AsIs("device_db.device")})
        # print(cur.fetchone())
        query="""INSERT INTO device_db.device (id,account_id,serial_number,folder_id,cluster_id,added,connected,friendly_name,mfg_version,current_version,short_model_name,long_model_name,mac_address,ip_address,logging,managed,tpm_binding_key,tpm_signing_key,start_date,end_date,data_retention,last_update,inactive,data_retention_license,firebox_to_replace,wgc_timestamp) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s, %s,%s,%s,%s,%s,%s,%s,%s,%s,%s) """
       ##mac_address,ip_address,logging,managed,tpm_binding_key,tpm_signing_key,start_date,end_date,data_retention,last_update,inactive,data_retention_license,firebox_to_replace,wgc_timestamp
        ## (%d,%d,%s,%d,%d,%d,%d,%s,%s,%s,%s,%s,%s,%d,%d, %d,%s,%s,%t,%t,%d,%t,%d,%d,%s)
        ## (2147483592, 2, 'FVE1000080037', 0, 0, 1, 1, 'YN Test Liberty', '12.3.1', '12.3.1', 'FireboxV-MC', 'FireboxV Extra Large', '8f:11:b0:90:9:7b', '172.16.236.183', 0, 0, 'f0000000', 'e0000000', datetime.datetime(2018, 10, 11, 12, 4, 19, 449000), datetime.datetime(2018, 11, 10, 12, 4, 19, 449000), 365, datetime.datetime(2020, 10, 19, 18, 13, 55, 648000), 0, 'a', None))
        # print(query)
        cur.execute(query,(6, 'WGC-1-ceded0650b574258838f', 'FVE1000080033', 0, 0, 1, 1, 'YN Test Liberty', '12.3.1', '12.3.1', 'FireboxV-MC', 'FireboxV Extra Large', '8f:11:b0:90:9:7b', '172.16.236.183', 0, 0, 'f0000000', 'e0000000', datetime.datetime(2018, 10, 11, 12, 4, 19, 449000), datetime.datetime(2018, 11, 10, 12, 4, 19, 449000), 365, datetime.datetime(2020, 10, 19, 18, 13, 55, 648000), 0, 0, "char",datetime.datetime(2020, 10, 19, 18, 13, 55, 648000)))
        cur.close()
        conn.commit()
        conn.close()
        # [4, 'ACC-TEST', '70AC028DE5013', nan, 0.0, 1, 0, 'T30', 11.4, nan, 'T30', 'Firebox T30', nan, nan, 1, 0, nan, nan, NaT, Timestamp('2022-04-01 12:34:56'), 60, Timestamp('2020-08-24 18:15:25.927000'), 0, 0, nan, Timestamp('2020-08-06 09:17:06')]
        # for r in range(1,10):
        #     # id=sheet.cell(r,0).value
        #     account_id=sheet.cell(r,1).value
        #     serial_number=sheet.cell(r,2).value
        #     folder_id=sheet.cell(r,3).value
        #     cluster_id=sheet.cell(r,4).value
        #     added=sheet.cell(r,5).value
        #     connected=sheet.cell(r,6).value
        #     friendly_name=sheet.cell(r,7).value
        #     mfg_version=sheet.cell(r,8).value
        #     current_version=sheet.cell(r,9).value
        #     short_model_name=sheet.cell(r,10).value
        #     long_model_name=sheet.cell(r,11).value
        #     mac_address=sheet.cell(r,12).value
        #     ip_address=sheet.cell(r,13).value
        #     logging=sheet.cell(r,14).value
        #     managed=sheet.cell(r,15).value
        #     tpm_binding_key=sheet.cell(r,16).value
        #     tpm_signing_key=sheet.cell(r,17).value
        #     start_date=sheet.cell(r,18).value
        #     end_date=sheet.cell(r,19).value
        #     data_retention=sheet.cell(r,20).value
        #     last_update=sheet.cell(r,21).value
        #     inactive=sheet.cell(r,22).value
        #     data_retention_license=sheet.cell(r,23).value
        #     firebox_to_replace=sheet.cell(r,24).value
        #     wgc_timestamp=sheet.cell(r,25).value
            
        #     values = (account_id,serial_number,folder_id,cluster_id,added,connected,friendly_name,mfg_version,current_version,short_model_name,long_model_name,mac_address,ip_address,logging,managed,tpm_binding_key,tpm_signing_key,start_date,end_date,data_retention,last_update,inactive,data_retention_license,firebox_to_replace,wgc_timestamp)
        #     print(f"{values=}")
        #     cur.execute(query,values)
        # cur.close()
        # conn.commit()
        # conn.close()

        # print(cur.fetchall())
    except Exception as e:
        print(e)
    return {
        "Connection established"
    }

def exceldata():
    pass
    # wb_obj = openpyxl.load_workbook(PATH)
    # sheet_obj = wb_obj.active
    # # cell_obj = sheet_obj.cell(row = 1, column = 1)
    # # print(cell_obj) 
    # cell_obj = sheet_obj['A1': 'X1']
    
    # for _ in cell_obj:
    #     print(_.value)
    # for row in sheet.rows:
    #     print(row[1].value)

# lambda_handler()
# exceldata()

df = pd.read_excel('devicedata.xlsx')

l1 = []

for index, row in df.iterrows():
    l1.append(row.to_list())

# print(l1)
lambda_handler()